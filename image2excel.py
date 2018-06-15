# -*- coding: utf-8 -*-
# 将图像转换为excel文件，一个像素点对应一个单元格
import re
import openpyxl
from PIL import Image
from openpyxl.styles import PatternFill


def tuple2str(t):
    lst = ['{0:#0{1}x}'.format(elt, 4) for elt in t]
    s = ''.join(lst)
    return re.sub('0x', '', s).upper()


def fill_color(s):
    color = PatternFill(fgColor=s, bgColor='00000000',fill_type='solid')
    return color

wb = openpyxl.Workbook()
ws = wb.active

im = Image.open('t.jpg')
for i in range(0,im.size[-1]):
    for j in range(0,im.size[0]):
        pixel = im.getpixel((j,i))
        pixel_str = tuple2str(pixel)
        print(i,j,pixel_str)
        color = fill_color(pixel_str)
        d = ws.cell(row=i+1, column=j+1)
        d.fill = color
#    if i == 1:
#        break

for col in ws.columns:
    column = col[0].column
    ws.column_dimensions[column].width=1

for row in ws.rows:
    print(row)
    row = row[0].row
    print(row)
    ws.row_dimensions[row].height=1
wb.save('out.xlsx')










